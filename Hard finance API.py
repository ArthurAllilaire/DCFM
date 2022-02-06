filepath = "DCFM.xlsx"
from openpyxl import *
import yahoo_fin.stock_info as si
from yahoo_fin import options
import pandas as pd
import numpy as np
ticker = "aapl"
wb = load_workbook(filepath)
# worksheetName = input("What is the worksheet name? ")
worksheetName = "Edited DCFM"
ws = wb[worksheetName]
"""
Debt = D13
Capex = D15
Cash = D14
I need to get the current year that is available for the numbers
in format DD/MM/YYYY
change in totalCurrentAssets - change in totalCurrentLiabilities - for last years
"""
def convertToNum(dataWithLetter):
  result = 0
  if "B" in dataWithLetter:
    dataWithoutLetter = dataWithLetter.replace("B","")
    result = float(dataWithoutLetter) * 1000000000
  elif "T" in dataWithLetter:
    dataWithoutLetter = dataWithLetter.replace("T","")
    result = float(dataWithoutLetter) * 1000000000000
  elif "M" in dataWithLetter:
    dataWithoutLetter = dataWithLetter.replace("M","")
    result = float(dataWithoutLetter) * 1000000
  return int(result)
def getCurrentData(data):
  years = data.keys()
  start = str(years[0])
  # start = start[:4]
  result = int(data[start])
  return result
def adder(data, row = 1, yearly = True, column = 4):
  if yearly:
    #Putting data into an array
    result = []
    years = data.keys()
    # print(years, type(years))
    start = str(years[-1])
    start = start[:4]
    # print(start)
    end = str(years[0])
    end = end[:4]
    for i in range(int(start),int(end)):
      year = str(i)
      result.append(int(data[year]))
    # Looping over array and putting it into excel
    r = 0
    for i in range(column,(column + len(result))):
      change = ws.cell(row,i)
      change.value = result[r]
      r += 1
  else:
    if isinstance(data, (int, float, str)):
      result = data
    else:
      years = data.keys()
      start = str(years[-1])
      # start = start[:4]
      result = int(data[start])
    change = ws.cell(row,column)
    change.value = result
def getYears(data):
  result = []
  years = data.keys()
  start = str(years[-1])
  startDate = start[:4]
  print("Start: " + startDate)
  end = str(years[0])
  endDate = end[:4]
  print("End: " + end)
  for i in range(int(startDate),int(endDate) + 1):
    year = str(i)
    result.append(int(data[year]))
    # print("works")
    # print(result)
  return result
def getYearsOnce(data):
  """
  Data should be panda series with dates, takes current date and changes it, then enters it into the excel sheet.
  """
  years = data.keys()
  for i in range(0,4):
    # Getting date
    end = str(years[i])
    # Reconfiguring year into right format
    endDate = end[8:10] + "/" + end[5:7] +"/" + end[:4]
    # if most recent date (i = 0) then 
    if i == 0:
    # ascribe it as fiscal year end
      ws["D10"] = endDate
    change = ws.cell(18,(9 - i))
    change.value = endDate
    # print(change)
financials = si.get_financials(ticker)
print(financials.keys())
incomeStatement = financials["yearly_income_statement"]
balanceSheet = financials["yearly_balance_sheet"]
cashFlow = financials["yearly_cash_flow"]
incomeStatementQuarterly = financials["quarterly_income_statement"]
balanceSheetQuarterly = financials["quarterly_balance_sheet"]
cashFlowQuarterly = financials["quarterly_cash_flow"]
print(incomeStatement, balanceSheet, cashFlow)
# Working down the table in excel
#tax rate = incomeTaxExpense/incomeBeforeTax * 100
incomeTaxExpense = getCurrentData(incomeStatementQuarterly.loc["incomeTaxExpense"])
incomeBeforeTax = getCurrentData(incomeStatementQuarterly.loc["incomeBeforeTax"])
TaxRate = (incomeTaxExpense/incomeBeforeTax) * 100
adder(TaxRate, row = 5, yearly = False)
#EBITDA
stats = si.get_stats(ticker)
EBITDA = convertToNum(stats.iloc[38]["Value"])
print(EBITDA)
print(stats)
totalCurrentLiabilities = balanceSheet.loc["totalCurrentLiabilities"]
getYearsOnce(totalCurrentLiabilities)
CLY = getYears(totalCurrentLiabilities)
print(CLY)
totalCurrentAssets = balanceSheet.loc["totalCurrentAssets"]
CLA = getYears(totalCurrentAssets)
print(CLA)
def getNWC(CLY, CLA):
  result = []
  NWC = []
  for i in range(0, len(CLY) - 1):
    # CLAChange = CLA[i] - CLA[i + 1]
    # CLYChange = CLY[i] - CLY[i + 1]
    workingCapital = CLA[i] - CLY[i]
    # result.append(CLAChange - CLYChange)
    result.append(workingCapital)
  print(result)
  for i in range(0, len(result) - 1):
    NWC.append(result[i] - result[i+1])
  return NWC
NWC = getNWC(CLY, CLA)
print(NWC)
wb.save(filepath)

from openpyxl import *
import yahoo_fin.stock_info as si
from yahoo_fin import options
import pandas as pd
import numpy as np
# Initialising the excel sheet
filepath = "Discounted-Cash-Flow-Model.xlsx"
wb = load_workbook(filepath)
# print(wb.sheetnames)
worksheetName = input("What is the worksheet name? ")
ws = wb[worksheetName]
ticker = str(ws["B12"].value)

def convertToNum(dataWithLetter):
  result = 0
  if "B" in dataWithLetter:
    dataWithoutLetter = dataWithLetter.replace("B","")
    result = float(dataWithoutLetter) * 1000000000
  elif "T" in dataWithLetter:
    dataWithoutLetter = dataWithLetter.replace("T","")
    result = float(dataWithoutLetter) * 1000000000000
  elif "M" in dataWithLetter:
    dataWithoutLetter = dataWithLetter.replace("M","")
    result = float(dataWithoutLetter) * 1000000
  return int(result)
def adder(data, row = 1, yearly = True, column = 2):
  if yearly:
    #Putting data into an array
    result = []
    years = data.keys()
    # print(years, type(years))
    start = str(years[-1])
    start = start[:4]
    # print(start)
    end = str(years[0])
    end = end[:4]
    for i in range(int(start),int(end)):
      year = str(i)
      result.append(int(data[year]))
    # Looping over array and putting it into excel
    r = 0
    for i in range(column,(column + len(result))):
      change = ws.cell(row,i)
      change.value = result[r]
      r += 1
  else:
    if isinstance(data, int) or isinstance(data, float):
      result = data
    elif isinstance(data, str):
      result = data
    else:
      years = data.keys()
      start = str(years[-1])
      start = start[:4]
      result = int(data[start])
    change = ws.cell(row,column)
    change.value = result

# historicStockPrice = si.get_data(ticker, start_date="01-01-2015")
# liveStockPrice = si.get_live_price(ticker)

financials = si.get_financials(ticker)
incomeStatement = financials["yearly_income_statement"]
balanceSheet = financials["yearly_balance_sheet"]
cashFlow = financials["yearly_cash_flow"]

# income statement data
totalRevenue = incomeStatement.loc["totalRevenue"]
adder(totalRevenue, row = 25)
interestExpense = incomeStatement.loc["interestExpense"]
adder(interestExpense, row = 26)
incomeBeforeTax = incomeStatement.loc["incomeBeforeTax"]
adder(incomeBeforeTax, row = 27)
incomeTaxExpense = incomeStatement.loc["incomeTaxExpense"]
adder(incomeTaxExpense, row = 28)
netIncome = incomeStatement.loc["netIncome"]
adder(netIncome, row = 29)

# cash flow to equity
cashFlowFromOperations = cashFlow.loc["totalCashFromOperatingActivities"]
adder(cashFlowFromOperations, row = 19)
capitalExpenditure = cashFlow.loc["capitalExpenditures"]
adder(capitalExpenditure, row = 20)
netBorrowings = cashFlow.loc["netBorrowings"]
adder(netBorrowings, row = 21)
# balance sheet stuff (WACC, CAPM)
balKeys = balanceSheet.keys()
if "shortLongTermDebt" in balKeys:
  shortLongTermDebt = balanceSheet.loc["shortLongTermDebt"]
else:
  shortLongTermDebt = 0
adder(shortLongTermDebt, row = 44, yearly = False)
longTermDebt = balanceSheet.loc["longTermDebt"]
adder(longTermDebt, row = 45, yearly = False)

#quote table
quote = si.get_quote_table(ticker)
marketCap = convertToNum(quote["Market Cap"])
adder(marketCap, row = 54, yearly = False)
beta = quote["Beta (5Y Monthly)"]
print(beta)
adder(beta, row = 52, yearly = False)

# analyst revenues
analysts = si.get_analysts_info(ticker)
revenueEstimatekeys = analysts["Revenue Estimate"].keys()
rightKey = str(revenueEstimatekeys[-2])
currentYear = int(rightKey[-5:-1])
ws["B11"] = currentYear
revenueCurrentYear = convertToNum(analysts["Revenue Estimate"].loc[1].loc[rightKey])
# Add analyst revenues
ws["F35"] = revenueCurrentYear
#Stats
stats = si.get_stats(ticker)
totalDebt = convertToNum(stats.iloc[44]["Value"])
adder(totalDebt, row = 55, yearly = False)
sharesOutstanding = convertToNum(stats.iloc[9]["Value"])
adder(sharesOutstanding, row = 14, yearly = False)



wb.save(filepath)

# endDate                         2019-12-31    2018-12-31    2017-12-31    2016-12-31
# Breakdown
# intangibleAssets              1.470335e+10  1.495114e+10  1.037106e+10  7.274501e+09
# totalLiab                     2.639356e+10  2.073564e+10  1.543079e+10  1.090681e+10
# totalStockholderEquity        7.582157e+09  5.238765e+09  3.581956e+09  2.679800e+09
# otherCurrentLiab              5.338306e+09  5.442461e+09  4.791663e+09  4.076183e+09
# totalAssets                   3.397571e+10  2.597440e+10  1.901274e+10  1.358661e+10
# commonStock                   2.793929e+09  2.315988e+09  1.871396e+09  1.599762e+09
# retainedEarnings              4.811749e+09  2.942359e+09  1.731117e+09  1.128603e+09
# otherLiab                     3.355987e+09  3.888257e+09  3.465042e+09  2.955842e+09
# treasuryStock                -2.352100e+07 -1.958200e+07 -2.055700e+07 -4.856500e+07
# otherAssets                   1.099664e+10  9.108430e+08  6.523090e+08  3.399060e+08
# cash                          5.018437e+09  3.794483e+09  2.822795e+09  1.467576e+09
# totalCurrentLiabilities       6.855696e+09  6.487320e+09  5.466312e+09  4.586657e+09
# deferredLongTermAssetCharges  6.582060e+08  5.644250e+08  4.782660e+08  2.272480e+08
# otherStockholderEquity       -2.352100e+07 -1.958200e+07 -2.055700e+07 -4.856500e+07
# propertyPlantEquipment        2.097221e+09  4.182810e+08  3.194040e+08  2.503950e+08
# totalCurrentAssets            6.178504e+09  9.694135e+09  7.669974e+09  5.720291e+09
# netTangibleAssets            -7.121195e+09 -9.712376e+09 -6.789099e+09 -4.594701e+09
# netReceivables                9.790680e+08  5.696330e+08           NaN           NaN
# longTermDebt                  1.475926e+10  1.036006e+10  6.499432e+09  3.364311e+09
# accountsPayable               6.743470e+08  5.629850e+08  3.595550e+08  3.128420e+08
# otherCurrentAssets                     NaN  5.151186e+09  4.847179e+09  3.986509e+09
# longTermInvestments                    NaN           NaN           NaN  1.517000e+06
# shortTermInvestments                   NaN           NaN           NaN  2.662060e+08